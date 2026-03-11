from flask import Blueprint, request, send_file
import os
import tempfile
import openpyxl
import random
import string

excel_process_bp = Blueprint('excel_process', __name__)



def adjust_prices(ws, target_total, limit_price=False, price_min=50, price_max=95):
    """调整价格以达到目标总价"""
    # 读取所有行的数据
    rows = []
    total_rows = 0
    
    # 假设从第2行开始有数据（第1行是表头）
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # 假设第一列是序号，非空表示有效行
            rows.append(list(row))
            total_rows += 1
    
    if not rows:
        return
    
    # 设定价格调整范围
    min_ratio = price_min / 100  # 转换为小数比例
    max_ratio = price_max / 100  # 转换为小数比例
    
    # 第一步：为每个物品生成随机价格
    for row in rows:
        # 假设第5列是单价，第6列是新单价
        original_price = row[4] if row[4] else 0
        if original_price > 0:
            # 生成随机比例（每个物品独立生成）
            ratio = random.uniform(min_ratio, max_ratio)
            new_price = original_price * ratio
            
            # 根据原价处理小数
            if original_price >= 100:
                new_price = int(round(new_price))  # 100以上的价格取整为整数
            else:
                new_price = round(new_price, 2)  # 100以下的价格保留两位小数
            
            row[5] = new_price
        else:
            row[5] = 0
    
    # 第二步：计算当前总价并调整
    def calculate_total():
        total = 0
        for row in rows:
            new_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            total += new_price * quantity
        return total
    
    current_total = calculate_total()
    
    # 第三步：调整价格使总价接近目标总价
    # 计算允许的误差范围（1%）
    error_threshold = target_total * 0.01
    
    # 循环调整直到误差在范围内
    while abs(current_total - target_total) > error_threshold:
        # 计算每个物品的总价
        item_totals = []
        for i, row in enumerate(rows):
            new_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            item_total = new_price * quantity
            item_totals.append((item_total, i))
        
        # 按总价从高到低排序
        item_totals.sort(reverse=True, key=lambda x: x[0])
        
        # 计算需要调整的金额
        adjustment_needed = target_total - current_total
        
        # 逐个调整物品价格
        for item_total, index in item_totals:
            if abs(adjustment_needed) < error_threshold:
                break
            
            row = rows[index]
            original_price = row[4] if row[4] else 0
            current_price = row[5] if row[5] else 0
            quantity = row[6] if row[6] else 0
            
            if quantity == 0:
                continue
            
            # 计算单个物品需要调整的金额
            item_adjustment = adjustment_needed / quantity
            
            # 计算新的价格
            new_price = current_price + item_adjustment
            
            # 确保价格在允许范围内
            min_price = original_price * min_ratio
            max_price = original_price * max_ratio
            
            # 根据原价处理小数
            if original_price >= 100:
                new_price = int(round(new_price))  # 100以上的价格取整为整数
                min_price = int(round(min_price))
                max_price = int(round(max_price))
            else:
                new_price = round(new_price, 2)  # 100以下的价格保留两位小数
                min_price = round(min_price, 2)
                max_price = round(max_price, 2)
            
            # 确保价格在范围内
            new_price = max(min_price, min(max_price, new_price))
            
            # 更新价格
            old_price = row[5]
            row[5] = new_price
            
            # 更新调整金额
            adjustment_needed -= (new_price - old_price) * quantity
            
        # 重新计算总价
        current_total = calculate_total()
    
    # 将调整后的数据写回Excel
    for i, row in enumerate(rows, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)
    
    # 计算最终总价
    final_total = calculate_total()
    print(f"目标总价: {target_total}, 调整后总价: {final_total}, 误差: {abs(final_total - target_total):.2f}")

@excel_process_bp.route('/process-excel', methods=['POST'])
def process_excel():
    """处理Excel文件，生成随机型号并调整价格"""
    try:
        # 检查是否有文件部分
        if 'file' not in request.files:
            return "没有选择文件", 400
        file = request.files['file']

        # 如果用户没有选择文件
        if file.filename == '':
            return "没有选择文件", 400

        # 检查文件类型
        if not file.filename.lower().endswith('.xlsx'):
            return "不支持的文件类型。请上传 .xlsx 文件。", 400

        # 获取目标总价
        target_total = request.form.get('targetTotalPrice', type=float)
        if target_total is None:
            return "请输入目标总价", 400
        
        # 获取是否限价
        limit_price = request.form.get('limitPrice', 'false').lower() == 'true'
        
        # 获取价格浮动区间
        price_min = request.form.get('priceMin', type=int, default=50)
        price_max = request.form.get('priceMax', type=int, default=95)

        # 使用临时文件处理
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file_path = temp_file.name
        temp_file.close()  # 关闭文件以便后续读取

        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(temp_file_path)
            ws = wb.active

            # 调整价格
            adjust_prices(ws, target_total, limit_price, price_min, price_max)

            # 保存处理后的文件
            output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            output_path = output_file.name
            output_file.close()
            wb.save(output_path)

            # 生成文件名
            original_name = os.path.basename(file.filename)
            new_filename = f"NEW_{original_name}"

            # 发送文件
            return send_file(output_path, as_attachment=True, download_name=new_filename)

        finally:
            # 清理临时文件
            if os.path.exists(temp_file_path):
                os.unlink(temp_file_path)

    except Exception as e:
        return f"处理文件时出错: {str(e)}", 500